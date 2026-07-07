"""
process_invoices_reference.py
------------------------------
Reference / test implementation of the invoice-processing logic described
in Main.xaml. This is NOT a replacement for the UiPath bot -- it mirrors
the same extraction, validation, and reporting steps in Python so the
logic can be verified and demoed without opening UiPath Studio.

Run:
    python3 process_invoices_reference.py

Reads:  sample_invoices/*.pdf
Writes: output/Invoice_Report.xlsx
Prints: automation summary (same shape as the bot's console/log output)
"""
import os
import re
import glob
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "sample_invoices")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Invoice_Report.xlsx")

# Same regex patterns as used in the UiPath "Extract Required Fields" step
PATTERNS = {
    "InvoiceNumber": r"Invoice Number\s*:\s*(.*)",
    "InvoiceDate":   r"Invoice Date\s*:\s*(.*)",
    "VendorName":    r"Vendor\s*:\s*(.*)",
    "Amount":        r"Amount\s*:\s*(?:Rs\.?|₹)?\s*([0-9,\.]+)",
}


def log(message):
    print(f"[LOG] {message}")


def read_invoice_text(pdf_path):
    """Equivalent of the 'Read PDF Text' activity."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_fields(invoice_text):
    """Equivalent of the 'Extract Required Fields' step (regex matching)."""
    fields = {}
    for field, pattern in PATTERNS.items():
        match = re.search(pattern, invoice_text)
        fields[field] = match.group(1).strip() if match else ""
    return fields


def validate_fields(fields):
    """Equivalent of the 'Validation' step."""
    if not fields["InvoiceNumber"]:
        return "Missing Invoice Number"
    if not fields["InvoiceDate"]:
        return "Missing Date"
    if not fields["VendorName"]:
        return "Missing Vendor"
    if not fields["Amount"]:
        return "Missing Amount"
    return "Valid"


def process_all_invoices():
    log("Process Started")
    rows = []
    pdf_files = sorted(glob.glob(os.path.join(INPUT_DIR, "*.pdf")))

    if not pdf_files:
        log(f"No invoices found in {INPUT_DIR}")
        return rows

    for pdf_path in pdf_files:
        name = os.path.basename(pdf_path)
        log(f"Reading {name}")
        text = read_invoice_text(pdf_path)

        log(f"Extracting Fields from {name}")
        fields = extract_fields(text)

        status = validate_fields(fields)
        log(f"Validation Result for {name}: {status}")

        rows.append({
            "Invoice Number": fields["InvoiceNumber"] or "(missing)",
            "Date": fields["InvoiceDate"],
            "Vendor": fields["VendorName"],
            "Amount": fields["Amount"],
            "Status": status,
        })

    return rows


def write_excel_report(rows):
    """Equivalent of the 'Write Excel' step."""
    log("Writing Excel Report")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    headers = ["Invoice Number", "Date", "Vendor", "Amount", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    valid_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in rows:
        ws.append([row["Invoice Number"], row["Date"], row["Vendor"], row["Amount"], row["Status"]])
        fill = valid_fill if row["Status"] == "Valid" else error_fill
        for col_idx in range(1, 6):
            ws.cell(row=ws.max_row, column=col_idx).fill = fill

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(header)] + [len(str(r.get(header, ""))) for r in rows] + [12])
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    wb.save(OUTPUT_FILE)
    log(f"Excel report saved to {OUTPUT_FILE}")


def print_summary(rows):
    """Equivalent of the 'Generate Summary' step."""
    total = len(rows)
    valid = sum(1 for r in rows if r["Status"] == "Valid")
    errors = total - valid

    summary = (
        "\nAutomation Summary\n"
        "-------------------\n"
        f"Total Invoices : {total}\n"
        f"Processed      : {total}\n"
        f"Valid          : {valid}\n"
        f"Errors         : {errors}\n"
    )
    print(summary)
    log("Process Completed")
    return {"total": total, "processed": total, "valid": valid, "errors": errors}


if __name__ == "__main__":
    invoice_rows = process_all_invoices()
    write_excel_report(invoice_rows)
    print_summary(invoice_rows)
